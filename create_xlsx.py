#!/usr/bin/env python3
import csv
import sys
sys.path.insert(0, '.')

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    # Create Test Cases XLSX
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"
    
    with open('02_Test_Cases/Search_Test_Cases.csv', 'r') as f:
        for row_idx, row in enumerate(csv.reader(f), 1):
            for col_idx, val in enumerate(row, 1):
                ws.cell(row_idx, col_idx, val)
    
    # Format header
    for col in range(1, 10):
        cell = ws.cell(1, col)
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    
    ws.freeze_panes = "A2"
    wb.save('02_Test_Cases/Search_Test_Cases.xlsx')
    print("Test Cases XLSX created")
    
    # Create Defect Log XLSX
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "Defects"
    
    with open('04_Defect_Log/Defect_Log.csv', 'r') as f:
        for row_idx, row in enumerate(csv.reader(f), 1):
            for col_idx, val in enumerate(row, 1):
                ws2.cell(row_idx, col_idx, val)
    
    for col in range(1, 12):
        cell = ws2.cell(1, col)
        cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    ws2.freeze_panes = "A2"
    wb2.save('04_Defect_Log/Defect_Log.xlsx')
    print("Defect Log XLSX created")
    
except ImportError:
    print("openpyxl not available, creating text-based alternatives")
